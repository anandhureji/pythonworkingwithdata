import openpyxl

# Open the Excel file
wb = openpyxl.load_workbook("pythonexcel.xlsx")

# Select the active sheet
sheet = wb.active

# Filter the data to include only rows where Age is greater than 30
filtered_data = [["Name", "Age", "City", "Amount"]]
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[1] > 30:
        filtered_data.append(list(row))

# Create a new Excel workbook and sheet
wb2 = openpyxl.Workbook()
sheet2 = wb2.active

# Write the filtered data to the sheet
for row in filtered_data:
    sheet2.append(row)

# Save the workbook to a new Excel file
wb2.save("output2.xlsx")
